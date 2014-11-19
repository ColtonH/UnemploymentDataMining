# Author: M.A. Menarguez

from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import os
from data.models import UsState,UnemploymentByStateMonthly

def getFilesInDir(path,ext):
    files=[]
    for f in os.listdir(path):
        if f.endswith("."+str(ext)):
            files.append(f)
    return files

def retrieve_data(path,files):
    data_aggregated = {}
    for f in files:
        wb = load_workbook(path+'/'+f)
        sheet_name = wb.get_sheet_names()[0]
        sheet_ranges = wb[sheet_name]
        state = sheet_ranges['B6'].value
        data= {}
        for row in range(12,51):
            year=int(sheet_ranges['A'+str(row)].value)
            data[year]=[]
            for col in range(2,14):
                data[year].append(sheet_ranges[get_column_letter(col)+str(row)].value)
        data_aggregated[state] = data
    return data_aggregated



def insert_unemployment_state(state,state_data):
    # Check if state exists
    print "Inserting "+ state
    try:
        state = UsState.objects.get(name=state)
    except Exception, e:
        # State does not exit, need to add it to the list
        state = UsState(name=state)
        try:
            state.save()
        except Exception, e:
            # Unhandled error
            print "Error adding state: "+state
            print "Exit..." 
            sys.exit(0)
    # Add all data
    for year in state_data.keys():
        # print year
        for month in range(0,len(state_data[year])):
            if state_data[year][month]:
                unemp_data_point = UnemploymentByStateMonthly(
                    state=state,
                    year=int(year),
                    month=month+1,
                    value=state_data[year][month]
                )
                unemp_data_point.save()

def insert_unemployment_all_states( data):
    for state in data.keys():
        insert_unemployment_state(state,data[state])

def main():
    path = '/home/menarguez/codes/unemployment'
    ext = 'xlsx'
    files = getFilesInDir(path,ext)
    print "Reading data"
    data = retrieve_data(path,files)
    # print(data)
    print "inserting data to db"
    insert_unemployment_all_states(data)

# if __name__ == "__main__":
main()
