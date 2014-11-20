# Author: M.A. Menarguez
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import os
from data.models import UsState,UnemploymentByStateMonthly,Race,NatalityByStateYearly

def getFilesInDir(path,ext):
    files=[]
    for f in os.listdir(path):
        if f.endswith("."+str(ext)):
            files.append(f)
    return files

def retrieve_data(path,files):
    data = {}
    for f in files:
        wb = load_workbook(path+'/'+f)
        sheet_name = wb.get_sheet_names()[0]
        sheet_ranges = wb[sheet_name]
        row = 1;
        while (True):
            row += 1
            state = sheet_ranges['B'+str(row)].value
            if (not state or state==""):
                break
            year = int(sheet_ranges['F'+str(row)].value)
            race =  sheet_ranges['D'+str(row)].value
            births =  int(sheet_ranges['H'+str(row)].value)
            birth_rate = sheet_ranges['J'+str(row)]
            if (birth_rate.value):
                birth_rate = float(birth_rate.value)
            else:
                birth_rate = None
            fertility_rate = sheet_ranges['L'+str(row)]
            if (fertility_rate.value):
                fertility_rate = float(fertility_rate.value)
            else:
                fertility_rate = None
            total_population = sheet_ranges['I'+str(row)]
            if (total_population.value):
                total_population = float(total_population.value)
            else:
                total_population = None
            if state not in data.keys():
                data[state]={}
            if year not in data[state].keys():
                data[state][year]={}
            data[state][year][race] = {
                'births': births,
                'birth_rate' : birth_rate,
                'total_population': total_population,
                'fertility_rate':fertility_rate,
            }
    return data

def get_race_or_add_it(race):
    try:
        race = Race.objects.get(name=race)
    except Exception, e:
        # State does not exit, need to add it to the list
        race = Race(name=race)
        try:
            race.save()
        except Exception, e:
            # Unhandled error
            print "Error adding race: "+race
            print "Exit..." 
            sys.exit(0)
    return race

def get_state_or_add_it(state):
    try:
        state = UsState.objects.get(name=state)
    except Exception, e:
        # State does not exit, need to add it to the list
        state = UsState(name=state)
        try:
            print "State %s is not available in the list of states" % (state)
            # state.save()
        except Exception, e:
            # Unhandled error
            print "Error adding state: "+state
            print "Exit..." 
            sys.exit(0)
    return state

def insert_natality_state(state,state_data):
    # Check if state exists
    print "Inserting "+ state
    state = get_state_or_add_it(state)
    # Add all data
    for year in state_data.keys():
        for race in state_data[year].keys():
            data = state_data[year][race]
            # print '%d: %s %d %s %s %s' % (year,race,data['births'],str(data['birth_rate']),str(data['total_population'],str(data['fertility_rate'])))
            race_model = get_race_or_add_it(race)
            if state_data[year][race]:
                natality_data_record = NatalityByStateYearly(
                    state=state,
                    race=race_model,
                    year=int(year),
                    num_births=data['births'],
                    birth_rate=data['birth_rate'],  
                    total_population=data['total_population'],
                    fertility_rate=data['fertility_rate'],
                )
                natality_data_record.save()

def insert_natality_all_states( data):
    #First delete all data (we are going to innsert it again)
    NatalityByStateYearly.objects.all().delete()
    for state in data.keys():
        insert_natality_state(state,data[state])

def main():
    path = '/webapps/unemployment_mining/unemployment_mining/data_import_and_extra_code/natality'
    ext = 'xlsx'
    files = getFilesInDir(path,ext)
    print "Reading data"
    data = retrieve_data(path,files)
    # print(data)
    # print(data)
    print "inserting data to db"
    insert_natality_all_states(data)

# if __name__ == "__main__":
main()
